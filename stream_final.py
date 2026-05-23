import openpyxl
import os
import random
import time
from datetime import datetime

path = r'C:\Users\Hisana jannath\OneDrive\Documents\LiveData.xlsx'

print("=" * 50)
print("  CODTECH TASK 3 - Real-Time Data Streamer")
print("  Simulated API Feed - Updates every 5 sec")
print("=" * 50)
print(f"  Writing to: {path}")
print("  Press Ctrl+C to stop\n")

count = 0
data_gb = 3.0

# Always create a fresh clean file
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'LiveData'
ws.append(['Timestamp', 'Active_Users', 'Requests_Per_Sec', 'Latency_ms', 'Error_Rate_Pct', 'Data_Ingested_GB'])

# Add 20 rows of historical data first
for i in range(20):
    from datetime import timedelta
    ts = datetime.now() - timedelta(seconds=(20-i)*5)
    data_gb += round(random.uniform(0.01, 0.05), 3)
    ws.append([
        ts.strftime('%Y-%m-%d %H:%M:%S'),
        random.randint(950, 2100),
        random.randint(580, 1150),
        round(random.uniform(18, 130), 1),
        round(random.uniform(0.1, 3.2), 2),
        round(data_gb, 2)
    ])

wb.save(path)
print("  Initial data written. Now streaming...\n")

while True:
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb['LiveData']

        # Keep max 50 rows
        if ws.max_row > 51:
            ws.delete_rows(2, ws.max_row - 51)

        data_gb += round(random.uniform(0.01, 0.05), 3)
        row = [
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            random.randint(950, 2100),
            random.randint(580, 1150),
            round(random.uniform(18, 130), 1),
            round(random.uniform(0.1, 3.2), 2),
            round(data_gb, 2)
        ]
        ws.append(row)
        wb.save(path)

        count += 1
        print(f"  [{count}] {row[0]} | Users: {row[1]} | RPS: {row[2]} | Latency: {row[3]}ms | Error: {row[4]}%")

        time.sleep(5)

    except KeyboardInterrupt:
        print("\n  Stopped.")
        break
    except Exception as e:
        print(f"  Retrying... ({e})")
        time.sleep(3)
